#!/usr/bin/env python3
"""Generate Simplified Chinese i18n assets for Wingsearch.

Bird common names are resolved from Wikidata by scientific name when possible.
Remaining card text is machine-translated and post-processed with a Wingspan
glossary. Icon markers such as [egg] are protected during translation.
"""

from __future__ import annotations

import json
import math
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw


ROOT = Path(__file__).resolve().parents[1]
SRC_DATA = ROOT / "src" / "assets" / "data"
SRC_I18N = SRC_DATA / "i18n"
DOCS_I18N = ROOT / "docs" / "assets" / "data" / "i18n"
SRC_ICONS = ROOT / "src" / "assets" / "icons" / "i18n"
DOCS_ICONS = ROOT / "docs" / "assets" / "icons" / "i18n"
CACHE_DIR = Path(os.environ.get("WINGSEARCH_ZH_CACHE", "/private/tmp/wingsearch-zh-i18n"))
TRANSLATION_CACHE_FILE = CACHE_DIR / "translations.json"
WIKIDATA_CACHE_FILE = CACHE_DIR / "wikidata-labels.json"

GOOGLE_TRANSLATE_URL = "https://translate.google.com/translate_a/single"
WIKIDATA_SPARQL_URL = "https://query.wikidata.org/sparql"
TRANSLATE_SEPARATOR = "<xsep/>"

ICON_RE = re.compile(r"\[([a-z][a-z\-_]*)\]")
CARD_LINK_RE = re.compile(r'(<strong applink="/card/(\d+)">)(.*?)(</strong>)')


BONUS_NAME_TRANSLATIONS = {
    "Anatomist": "解剖学家",
    "Apiarist": "养蜂人",
    "Avian Theriogenologist": "鸟类繁殖专家",
    "Backyard Birder": "后院观鸟者",
    "Behaviorist": "行为学家",
    "Bird Bander": "鸟类环志员",
    "Bird Counter": "鸟类计数员",
    "Bird Feeder": "喂鸟人",
    "Breeding Manager": "繁殖管理者",
    "Brilliant Specialist": "辉煌专家",
    "Cartographer": "制图师",
    "Citizen Scientist": "公民科学家",
    "Diet Specialist": "饮食专家",
    "Ecolodge Owner": "生态旅馆业主",
    "Ecologist": "生态学家",
    "Emerald Specialist": "绿宝石专家",
    "Enclosure Builder": "围栏建造者",
    "Endangered Species Protector": "濒危物种保护者",
    "Ethologist": "动物行为学家",
    "Falconer": "驯隼师",
    "Fishery Manager": "渔业管理者",
    "Food Web Expert": "食物网专家",
    "Forest Data Analyst": "森林数据分析师",
    "Forest Population Monitor": "森林种群监测员",
    "Forest Ranger": "森林巡护员",
    "Forester": "林务员",
    "Grassland Data Analyst": "草原数据分析师",
    "Grassland Population Monitor": "草原种群监测员",
    "Grassland Ranger": "草原巡护员",
    "Historian": "历史学家",
    "Hummingbird Counter": "蜂鸟计数员",
    "Hummingbird Gardener": "蜂鸟园丁",
    "Large Bird Specialist": "大型鸟类专家",
    "Mango Specialist": "芒果专家",
    "Mechanical Engineer": "机械工程师",
    "Nest Box Builder": "巢箱建造者",
    "Omnivore Specialist": "杂食专家",
    "Oologist": "鸟卵学家",
    "Passerine Specialist": "雀形鸟专家",
    "Pellet Dissector": "食丸剖析者",
    "Photographer": "摄影师",
    "Platform Builder": "平台巢建造者",
    "Prairie Manager": "草原管理者",
    "Rodentologist": "啮齿动物学家",
    "Site Selection Expert": "选址专家",
    "Small Clutch Specialist": "小窝卵数专家",
    "Topaz Specialist": "黄玉专家",
    "Visionary Leader": "远见领袖",
    "Viticulturalist": "葡萄栽培师",
    "Wetland Data Analyst": "湿地数据分析师",
    "Wetland Population Monitor": "湿地种群监测员",
    "Wetland Ranger": "湿地巡护员",
    "Wetland Scientist": "湿地科学家",
    "Wildlife Gardener": "野生动物园丁",
    "Winter Feeder": "冬季喂食者",
    "[Fan Made] Caprimulgiform Specialist": "[粉丝自制] 夜鹰目与雨燕目专家",
    "[automa] Autwitcher": "[automa] 自动观鸟者",
    "[automa] Avid Asian Avian Admirer": "[automa] 热衷亚洲鸟类欣赏者",
    "[automa] Charm Champion": "[automa] 魅力冠军",
    "[automa] RASPB Life Fellow": "[automa] RASPB 终身会员",
    "[automa] Rare Species Lister": "[automa] 稀有物种记录者",
}

OTHER_TRANSLATIONS = {
    "WHEN ACTIVATED": "激活时",
    "WHEN PLAYED": "打出时",
    "ONCE BETWEEN TURNS": "两回合间一次",
    "ROUND END": "本轮结束时",
    "GAME END": "游戏结束时",
    "of cards": "的卡牌",
    "Additional Asian Avians": "额外亚洲鸟类",
    "Birds of Canada / Oiseaux du Canada": "加拿大鸟类 / Oiseaux du Canada",
    "Birds of Continental Europe / Ngā Manu o Aotearoa": "欧洲大陆鸟类 / Ngā Manu o Aotearoa",
    "Birds of New Zealand": "新西兰鸟类",
    "Birds of U.S.A.": "美国鸟类",
    "British Birds": "英国鸟类",
}

TERM_REPLACEMENTS = [
    ("birdfeeder", "喂食器"),
    ("Birdfeeder", "喂食器"),
    ("喂鸟器", "喂食器"),
    ("鸟食台", "喂食器"),
    ("喂鸟台", "喂食器"),
    ("鸟类喂食器", "喂食器"),
    ("鸟喂食器", "喂食器"),
    ("从喂食器处", "从喂食器中"),
    ("鸟箱", "鸟巢箱"),
    ("鸟屋", "鸟巢箱"),
    ("缓存", "贮藏"),
    ("被贮藏", "被贮藏"),
    ("藏匿", "贮藏"),
    ("藏在", "贮藏在"),
    ("塞入卡片", "塞牌"),
    ("塞入牌", "塞牌"),
    ("塞入的牌", "塞入的卡牌"),
    ("塞在这只鸟下面", "塞到这只鸟下方"),
    ("塞到这只鸟下面", "塞到这只鸟下方"),
    ("垫在这只鸟下面", "塞到这只鸟下方"),
    ("电源", "能力"),
    ("力量颜色", "能力颜色"),
    ("捕食者能力", "捕食能力"),
    ("圆结束", "本轮结束"),
    ("回合结束目标", "轮末目标"),
    ("游戏结束时结束", "游戏结束时"),
    ("供应", "供应区"),
    ("供应区区", "供应区"),
    ("个人供应区", "个人供应区"),
    ("鸟类卡", "鸟卡"),
    ("鸟牌", "鸟卡"),
    ("卡片", "卡牌"),
    ("奖金卡", "奖励卡"),
    ("奖励牌", "奖励卡"),
    ("目标卡牌", "目标卡"),
    ("栖息地垫", "玩家面板"),
    ("玩家垫", "玩家面板"),
    ("自动机", "Automa"),
    ("奥托玛", "Automa"),
    ("蜂鸟轨道", "蜂鸟轨"),
    ("食物成本", "食物费用"),
    ("获胜点", "分"),
    ("胜利点", "分"),
    ("积分值", "分值"),
    ("绘制", "抽取"),
    ("甲板", "牌堆"),
    ("托盘", "展示区"),
    ("模具", "骰子"),
    ("小鸟", "鸟"),
    ("鸡蛋", "蛋"),
    ("扩张", "扩展"),
    ("通风口", "泄殖腔"),
    ("您", "你"),
]

GOAL_TERM_TRANSLATIONS = {
    "[bird]": "[bird]",
    "[forest]": "[forest]",
    "[grassland]": "[grassland]",
    "[wetland]": "[wetland]",
    "[egg]": "[egg]",
    "[bowl]": "[bowl]",
    "[cavity]": "[cavity]",
    "[ground]": "[ground]",
    "[platform]": "[platform]",
    "[card]": "[card]",
    "[wild]": "[wild]",
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_cache(path: Path) -> dict[str, str]:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def save_cache(path: Path, data: dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


translation_cache = load_cache(TRANSLATION_CACHE_FILE)
wikidata_cache = load_cache(WIKIDATA_CACHE_FILE)


def protect_icons(text: str) -> tuple[str, list[tuple[str, str]]]:
    replacements: list[tuple[str, str]] = []

    def replace(match: re.Match[str]) -> str:
        token = f"<xicon{len(replacements)}/>"
        replacements.append((token, match.group(0)))
        return token

    return ICON_RE.sub(replace, text), replacements


def restore_icons(text: str, replacements: list[tuple[str, str]]) -> str:
    for token, original in replacements:
        text = text.replace(token, original)
        text = text.replace(token.replace("/", " /"), original)
    return text


def normalize_text(text: str) -> str:
    text = text.replace(" / >", "/>")
    text = re.sub(r"\s+([,.;:!?])", r"\1", text)
    text = re.sub(r"([（(])\s+", r"\1", text)
    text = re.sub(r"\s+([）)])", r"\1", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = text.replace(" <br> ", "<br>").replace(" <br>", "<br>").replace("<br> ", "<br>")
    text = text.replace("< /", "</")
    text = text.replace("<strongapplink", "<strong applink")
    text = text.replace("<i >", "<i>").replace("</ i>", "</i>")
    return text


def postprocess(text: str) -> str:
    if not text:
        return text
    text = normalize_text(text)
    for src, dst in TERM_REPLACEMENTS:
        text = text.replace(src, dst)
    text = text.replace("喂食器器", "喂食器")
    text = text.replace("贮藏贮藏", "贮藏")
    text = text.replace("贮贮藏", "贮藏")
    text = text.replace("贮藏藏", "贮藏")
    text = text.replace("供应区区", "供应区")
    text = text.replace("展示区区", "展示区")
    text = re.sub(r"抓\s*([0-9]+)\s*张", r"抽取 \1 张", text)
    text = re.sub(r"抓\s*([0-9]+)\s*个", r"抽取 \1 个", text)
    text = text.replace("抓牌", "抽牌")
    text = text.replace("抓到", "抽到")
    text = re.sub(r"([0-9])\+\s*鸟", r"\1+ 只鸟", text)
    text = re.sub(r"([0-9])\s+鸟", r"\1 只鸟", text)
    text = text.replace("玩另一只鸟", "打出另一只鸟")
    text = text.replace("再玩另一只鸟", "再打出另一只鸟")
    text = text.replace("玩一只鸟", "打出一只鸟")
    text = text.replace("玩鸟", "打出鸟")
    text = text.replace("你可以免费玩其中一只鸟", "你可以免费打出其中一只鸟")
    text = text.replace("异能", "能力")
    text = text.replace("[automa]", "[automa]")
    return normalize_text(text)


def translate_raw(text: str, source_language: str = "en") -> str:
    if not text:
        return text
    cache_key = f"{source_language}|zh-CN|{text}"
    if cache_key in translation_cache:
        return translation_cache[cache_key]

    data = urllib.parse.urlencode({
        "client": "gtx",
        "sl": source_language,
        "tl": "zh-CN",
        "dt": "t",
        "q": text,
    }).encode("utf-8")
    request = urllib.request.Request(
        GOOGLE_TRANSLATE_URL,
        data=data,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    for attempt in range(5):
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                payload = json.loads(response.read().decode("utf-8"))
            translated = "".join(segment[0] for segment in payload[0] if segment[0])
            translation_cache[cache_key] = translated
            if len(translation_cache) % 50 == 0:
                save_cache(TRANSLATION_CACHE_FILE, translation_cache)
            time.sleep(0.08)
            return translated
        except urllib.error.HTTPError as error:
            if attempt == 4:
                raise
            wait = 45 * (attempt + 1) if error.code == 429 else 2 * (attempt + 1)
            print(f"Translate HTTP {error.code}; retrying in {wait}s", flush=True)
            time.sleep(wait)
        except Exception:
            if attempt == 4:
                raise
            time.sleep(1.5 * (attempt + 1))
    return text


def translate_joined_raw(texts: list[str], source_language: str = "en") -> list[str]:
    joined = f"\n{TRANSLATE_SEPARATOR}\n".join(texts)
    data = urllib.parse.urlencode({
        "client": "gtx",
        "sl": source_language,
        "tl": "zh-CN",
        "dt": "t",
        "q": joined,
    }).encode("utf-8")
    request = urllib.request.Request(
        GOOGLE_TRANSLATE_URL,
        data=data,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    for attempt in range(5):
        try:
            with urllib.request.urlopen(request, timeout=45) as response:
                payload = json.loads(response.read().decode("utf-8"))
            translated = "".join(segment[0] for segment in payload[0] if segment[0])
            parts = [part.strip() for part in re.split(rf"\s*{re.escape(TRANSLATE_SEPARATOR)}\s*", translated)]
            if len(parts) != len(texts):
                raise ValueError(f"separator split returned {len(parts)} parts for {len(texts)} inputs")
            time.sleep(0.8)
            return parts
        except urllib.error.HTTPError as error:
            if attempt == 4:
                raise
            wait = 60 * (attempt + 1) if error.code == 429 else 3 * (attempt + 1)
            print(f"Batch translate HTTP {error.code}; retrying in {wait}s", flush=True)
            time.sleep(wait)
        except Exception as error:
            if attempt == 4:
                raise
            print(f"Batch translate failed ({error}); retrying", flush=True)
            time.sleep(3 * (attempt + 1))
    return texts


def warm_translation_cache(texts: list[Any], source_language: str = "en") -> None:
    protected_texts: list[str] = []
    seen: set[str] = set()
    for value in texts:
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        protected, _ = protect_icons(text)
        cache_key = f"{source_language}|zh-CN|{protected}"
        if cache_key in translation_cache or protected in seen:
            continue
        seen.add(protected)
        protected_texts.append(protected)

    chunks: list[list[str]] = []
    current: list[str] = []
    current_size = 0
    for text in protected_texts:
        next_size = current_size + len(text) + len(TRANSLATE_SEPARATOR) + 2
        if current and (next_size > 4500 or len(current) >= 24):
            chunks.append(current)
            current = []
            current_size = 0
        current.append(text)
        current_size += len(text) + len(TRANSLATE_SEPARATOR) + 2
    if current:
        chunks.append(current)

    total = len(chunks)
    for index, chunk in enumerate(chunks, start=1):
        try:
            translated_parts = translate_joined_raw(chunk, source_language)
        except Exception as error:
            print(f"Batch {index}/{total} failed ({error}); falling back to individual requests", flush=True)
            translated_parts = [translate_raw(text, source_language) for text in chunk]
        for source, translated in zip(chunk, translated_parts):
            translation_cache[f"{source_language}|zh-CN|{source}"] = translated
        if index % 10 == 0 or index == total:
            save_cache(TRANSLATION_CACHE_FILE, translation_cache)
            print(f"Warmed translation cache {index}/{total} batches", flush=True)


def translate_text(text: Any, source_language: str = "en") -> str | None:
    if text is None:
        return None
    text = str(text).strip()
    if not text:
        return None
    protected, replacements = protect_icons(text)
    translated = translate_raw(protected, source_language)
    translated = restore_icons(translated, replacements)
    return postprocess(translated)


def resolve_wikidata_labels(scientific_names: list[str]) -> dict[str, str]:
    missing = [name for name in sorted(set(scientific_names)) if name and name not in wikidata_cache]
    for start in range(0, len(missing), 80):
        chunk = missing[start:start + 80]
        values = " ".join(f'"{name}"' for name in chunk)
        query = f"""
SELECT ?taxonName ?zhLabel WHERE {{
  VALUES ?taxonName {{ {values} }}
  ?item wdt:P225 ?taxonName.
  ?item rdfs:label ?zhLabel.
  FILTER(LANG(?zhLabel) IN ("zh-hans", "zh-cn", "zh-sg", "zh", "zh-hant", "zh-tw"))
}}
"""
        url = WIKIDATA_SPARQL_URL + "?" + urllib.parse.urlencode({"format": "json", "query": query})
        request = urllib.request.Request(
            url,
            headers={"User-Agent": "Wingsearch zh i18n generator/1.0 (https://navarog.github.io/wingsearch/)"},
        )
        with urllib.request.urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8"))

        candidates: dict[str, list[tuple[int, str]]] = {name: [] for name in chunk}
        language_rank = {
            "zh-hans": 0,
            "zh-cn": 1,
            "zh-sg": 2,
            "zh": 3,
            "zh-hant": 4,
            "zh-tw": 5,
        }
        for row in payload["results"]["bindings"]:
            taxon = row["taxonName"]["value"]
            label = row["zhLabel"]["value"]
            lang = row["zhLabel"].get("xml:lang", "zh")
            candidates.setdefault(taxon, []).append((language_rank.get(lang, 99), label))

        for name in chunk:
            labels = sorted(candidates.get(name, []))
            wikidata_cache[name] = labels[0][1] if labels else ""
        save_cache(WIKIDATA_CACHE_FILE, wikidata_cache)
        print(f"Resolved Wikidata labels {min(start + len(chunk), len(missing))}/{len(missing)}", flush=True)
        time.sleep(0.4)
    return dict(wikidata_cache)


def get_bird_name(card: dict[str, Any], wikidata_labels: dict[str, str]) -> str:
    scientific_name = card.get("Scientific name") or ""
    label = wikidata_labels.get(scientific_name, "")
    if label:
        if re.search(r"[\u9fff-\ufaff]", label):
            return postprocess(label)
        return postprocess(translate_text(label) or label)
    return translate_text(card["Common name"]) or card["Common name"]


def translate_goal_text(text: str) -> str:
    replacements = {
        "in one row": "在同一行",
        "with no [egg]": "没有[egg]",
        "worth >4 [feather]": "分值大于 4 [feather]",
        "birds with tucked cards": "有塞牌的鸟",
        "food cost of played [bird]": "已打出[bird]的食物费用",
        "with [egg]": "有[egg]",
        "in hand": "在手牌中",
        "in personal supply": "在个人供应区",
        "filled columns": "填满的列",
    }
    out = text
    for src, dst in GOAL_TERM_TRANSLATIONS.items():
        out = out.replace(src, dst)
    for src, dst in replacements.items():
        out = out.replace(src, dst)
    if out == text:
        out = translate_text(text) or text
    return postprocess(out)


def english_goal_label(text: str) -> str:
    replacements = {
        "[bird]": "Bird",
        "[forest]": "Forest",
        "[grassland]": "Grassland",
        "[wetland]": "Wetland",
        "[egg]": "Egg",
        "[bowl]": "Bowl Nest",
        "[cavity]": "Cavity Nest",
        "[ground]": "Ground Nest",
        "[platform]": "Platform Nest",
        "[card]": "Card",
        "[wild]": "Food",
        "[bird_with_tucked_card]": "Bird with tucked card",
        "[feather]": "Points",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text[:1].upper() + text[1:]


def replace_link_labels(text: str, card_names: dict[int, str], bonus_names: dict[int, str]) -> str:
    if not text:
        return text

    def replace(match: re.Match[str]) -> str:
        card_id = int(match.group(2))
        label = card_names.get(card_id) or bonus_names.get(card_id) or match.group(3)
        return f"{match.group(1)}{label}{match.group(4)}"

    return CARD_LINK_RE.sub(replace, text)


def workbook_write_sheet(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for col_index, header in enumerate(headers, start=1):
        values = [str(row[col_index - 1]) if row[col_index - 1] is not None else "" for row in rows[:200]]
        width = max(len(header), *(len(value) for value in values)) + 2
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max(width, 12), 48)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def save_workbook(path: Path, rows_by_sheet: dict[str, tuple[list[str], list[list[Any]]]]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for name, (headers, rows) in rows_by_sheet.items():
        sheet = workbook.create_sheet(name)
        workbook_write_sheet(sheet, headers, rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def star_points(cx: float, cy: float, outer: float, inner: float) -> list[tuple[float, float]]:
    points = []
    start_angle = -math.pi / 2
    for index in range(10):
        radius = outer if index % 2 == 0 else inner
        angle = start_angle + index * math.pi / 5
        points.append((cx + radius * math.cos(angle), cy + radius * math.sin(angle)))
    return points


def write_language_icons() -> None:
    image = Image.new("RGB", (160, 80), "#DE2910")
    draw = ImageDraw.Draw(image)
    draw.polygon(star_points(25, 22, 12, 5), fill="#FFDE00")
    small_stars = [(48, 10, -25), (59, 21, 10), (58, 37, 30), (45, 50, 55)]
    for cx, cy, rotation in small_stars:
        points = star_points(cx, cy, 5, 2.1)
        draw.polygon(points, fill="#FFDE00")

    for folder in (SRC_ICONS, DOCS_ICONS):
        folder.mkdir(parents=True, exist_ok=True)
        image.save(folder / "zh.jpg", quality=92)
        image.save(folder / "zh.webp", quality=92)


def main() -> None:
    birds = load_json(SRC_DATA / "master.json") + load_json(SRC_DATA / "hummingbirds.json")
    bonuses = load_json(SRC_DATA / "bonus.json")
    goals = load_json(SRC_DATA / "goals.json")

    wikidata_labels = resolve_wikidata_labels([card.get("Scientific name", "") for card in birds])
    texts_to_translate: list[Any] = []
    for card in birds:
        label = wikidata_labels.get(card.get("Scientific name", ""), "")
        if not label or not re.search(r"[\u9fff-\ufaff]", label):
            texts_to_translate.append(label or card.get("Common name"))
        texts_to_translate.extend([card.get("Power text"), card.get("Flavor text"), card.get("Note")])
    for card in bonuses:
        if card["Bonus card"] not in BONUS_NAME_TRANSLATIONS:
            texts_to_translate.append(card["Bonus card"])
        texts_to_translate.extend([card.get("Condition"), card.get("Explanatory text"), card.get("VP"), card.get("Note")])
    for card in goals:
        texts_to_translate.extend([card.get("Goal"), card.get("Reverse")])
    print("Warming translation cache", flush=True)
    warm_translation_cache(texts_to_translate)

    bird_name_by_id = {int(card["id"]): get_bird_name(card, wikidata_labels) for card in birds}
    bonus_name_by_id = {
        int(card["id"]): BONUS_NAME_TRANSLATIONS.get(card["Bonus card"], translate_text(card["Bonus card"]) or card["Bonus card"])
        for card in bonuses
    }

    zh_birds: dict[str, dict[str, Any]] = {}
    bird_rows: list[list[Any]] = []
    bonus_columns = ["Anatomist", "Cartographer", "Historian", "Photographer"]
    for card in birds:
        card_id = int(card["id"])
        translated = {
            "English name": card["Common name"],
            "Scientific name": card["Scientific name"],
            "Expansion": card["Set"],
            "Common name": bird_name_by_id[card_id],
            "Power text": translate_text(card.get("Power text")) if card.get("Power text") else None,
            "Flavor text": translate_text(card.get("Flavor text")) if card.get("Flavor text") else None,
            "Note": translate_text(card.get("Note")) if card.get("Note") else None,
            **{column: card.get(column) for column in bonus_columns},
        }
        translated["Note"] = replace_link_labels(translated["Note"], bird_name_by_id, bonus_name_by_id) if translated["Note"] else None
        zh_birds[str(card_id)] = translated
        bird_rows.append([
            card_id,
            translated["English name"],
            translated["Scientific name"],
            translated["Expansion"],
            translated["Common name"],
            translated["Power text"],
            translated["Flavor text"],
            translated["Note"],
            translated["Anatomist"],
            translated["Cartographer"],
            translated["Historian"],
            translated["Photographer"],
        ])

    zh_bonuses: dict[str, dict[str, Any]] = {}
    bonus_rows: list[list[Any]] = []
    for card in bonuses:
        card_id = int(card["id"])
        translated = {
            "English name": card["Bonus card"],
            "Expansion": card["Set"],
            "Name": bonus_name_by_id[card_id],
            "Condition": translate_text(card.get("Condition")) if card.get("Condition") else None,
            "Explanatory text": translate_text(card.get("Explanatory text")) if card.get("Explanatory text") else None,
            "VP": translate_text(card.get("VP")) if card.get("VP") else None,
            "Note": translate_text(card.get("Note")) if card.get("Note") else None,
        }
        for key in ("Condition", "Explanatory text", "VP", "Note"):
            if translated[key]:
                translated[key] = replace_link_labels(translated[key], bird_name_by_id, bonus_name_by_id)
        zh_bonuses[str(card_id)] = translated
        bonus_rows.append([
            card_id,
            translated["English name"],
            translated["Expansion"],
            translated["Name"],
            translated["Condition"],
            translated["Explanatory text"],
            translated["VP"],
            translated["Note"],
        ])

    zh_goals: dict[str, dict[str, Any]] = {}
    goal_rows: list[list[Any]] = []
    for card in goals:
        card_id = int(card["id"])
        translated = {
            "English name": english_goal_label(card["Goal"]),
            "Expansion": card["Set"],
            "Name": translate_goal_text(card["Goal"]),
            "Condition": translate_goal_text(card.get("Reverse", "")) if card.get("Reverse") else None,
            "Explanatory Text": None,
        }
        zh_goals[str(card_id)] = translated
        goal_rows.append([
            card_id,
            translated["English name"],
            translated["Expansion"],
            translated["Name"],
            translated["Condition"],
            translated["Explanatory Text"],
        ])

    zh_other = {key: {"Translated": value} for key, value in OTHER_TRANSLATIONS.items()}
    zh_parameters = {"Show bonus cards match symbols": {"Value": False}}

    result = {
        "birds": zh_birds,
        "bonuses": zh_bonuses,
        "goals": zh_goals,
        "other": zh_other,
        "parameters": zh_parameters,
    }

    save_json(SRC_I18N / "zh.json", result)
    save_json(DOCS_I18N / "zh.json", result)
    save_cache(TRANSLATION_CACHE_FILE, translation_cache)
    save_cache(WIKIDATA_CACHE_FILE, wikidata_cache)

    rows_by_sheet = {
        "Birds": (
            ["id", "English name", "Scientific name", "Expansion", "Common name", "Power text", "Flavor text", "Note", *bonus_columns],
            bird_rows,
        ),
        "Bonuses": (
            ["id", "English name", "Expansion", "Name", "Condition", "Explanatory text", "VP", "Note"],
            bonus_rows,
        ),
        "Goals": (
            ["id", "English name", "Expansion", "Name", "Condition", "Explanatory Text"],
            goal_rows,
        ),
        "Other": (
            ["English name", "Translated"],
            [[key, value["Translated"]] for key, value in zh_other.items()],
        ),
        "Parameters": (
            ["Name", "Value"],
            [[key, value["Value"]] for key, value in zh_parameters.items()],
        ),
    }
    save_workbook(ROOT / "i18n" / "zh.xlsx", rows_by_sheet)
    write_language_icons()
    print(f"Wrote {SRC_I18N / 'zh.json'}")
    print(f"Wrote {DOCS_I18N / 'zh.json'}")
    print(f"Wrote {ROOT / 'i18n' / 'zh.xlsx'}")


if __name__ == "__main__":
    main()
